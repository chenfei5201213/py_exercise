#-*- coding: utf-8 -*-

"""
Created on 2016年1月13日

@author: 守望@天空~
"""
import openpyxl
import ConfigParser
import os

class Excel(object):
    """
     Excel表格的测试数据驱动类
    """

    def __init__(self, ExcelPath):
        self.path = ExcelPath
        if not os.path.isfile(ExcelPath):  # 校验文件是否有效
            self.stat = False
            # "文件不存在，自动创建"
            self.wb = openpyxl.Workbook()
        else:
            self.stat = True
            self.wb = openpyxl.load_workbook(ExcelPath)


    def getData(self, *sheetnames):
        """
        获取整个excel的数据，每行都是dict
        :return: ｛sheetname:[{u'用例id':'1',"请求方法":'get',"url":"www.xx.com"},
        {u'用例id':'2',"请求方法":'get',"url":"www.xx.com"}],sheetname2:[]｝
        """
        if not self.stat:
            return {}
        if not sheetnames:
            sheetnames = self.wb.sheetnames
        all_data = {}
        for sheetname in sheetnames:
            sheet = self.wb[sheetname]
            # 按行列提取数据
            table_data = list(sheet.rows)
            titles = map(lambda a: a.value, table_data[0])
            data = table_data[1:]
            sheet_data = []
            # 按行组装字典
            for row in data:
                row_data = {}
                for i, title in enumerate(titles):
                    row_data[title] = row[i].value
                sheet_data.append(row_data)
            all_data[sheetname] = sheet_data
        return all_data

    def key_struct(self, key):
        # 根据指定的唯一值字段，生成结构字典
        all_data = {}
        for sheetname in self.wb.sheetnames:
            sheet = self.wb[sheetname]
            table_data = list(sheet.rows)
            titles = map(lambda a: a.value, table_data[0])
            if key not in titles:
                continue
            data = table_data[1:]
            sheet_data = {}
            for row in data:
                row_data = {}
                for i, title in enumerate(titles):
                    row_data[title] = row[i]
                sheet_data[row_data[key].value] = row_data
            all_data[sheetname] = sheet_data
        return all_data

    def update_data(self, data, key):
        if not self.stat:
            return False
        if not isinstance(data, dict):
            self._invaid_data('dict', data)
        table_struct = self.key_struct(key)
        for sheet_name, sheet_data in data.items():
            if sheet_name in table_struct:
                sheet0 = table_struct[sheet_name]
                for row in sheet_data:
                    row0 = sheet0[row[key]]
                    row.pop(key)
                    for keys, item in row.items():
                        row0[keys].value = item
        self.save()

    def add_sheet_data(self, data):
        for sheet_name in data:
            if sheet_name in self.wb.sheetnames:
                ws_data = zip(*map(lambda a: [a[0][0]]+list(a[1]), map(zip, *map(lambda a: a.items(), data[sheet_name]))))
                ws = self.wb[sheet_name]
                for row in ws_data[1:]:
                    ws.append(row)
            else:
                ws_data = zip(*map(lambda a: [a[0][0]] + list(a[1]), map(zip, *map(lambda a: a.items(), data[sheet_name]))))
                ws = self.wb.create_sheet(title=sheet_name)
                ws.column_dimensions["A"].width = 60.0
                for row in ws_data:
                    ws.append(row)
                self.stat = True
        self.save()

    def save(self, file_name=''):
        if not file_name:
            file_name = self.path
        self.wb.save(file_name)

    def _invaid_data(self, type_, data):
        raise TypeError('Value must be a {0}. Supplied value is {1}'.format(type_, type(data)))


class Config(dict):
    # 将config文件转为dict
    def __init__(self, configpath, **kwargs):
        super(Config, self).__init__(**kwargs)
        self.conf = ConfigParser.ConfigParser()
        self.conf.read(configpath)
        for section in self.conf.sections():
            options = {}
            for option in self.conf.options(section):
                value = self.conf.get(section, option)
                if not isinstance(value, unicode):
                    value = value.decode("utf-8")
                options[option] = value
            self[section] = options


if __name__ == "__main__":
    wb = Excel(u'测试一下.xlsx')
    data = {u'测试2': [{u'测试字段23': 1L, u'ID': 1L}, {u'测试字段23': 2L, u'ID': 2L}]}
    wb.add_sheet_data(data)
    data2={u'新闻': [{u'结果': u"测试通过",u"实际值": 233, u'用例ID': 1},
                    {u'结果': u"你瞅啥", u"实际值": 10086, u'用例ID': 2},
                    {u'结果': u"也过了",u"实际值": "", u'用例ID': 3}]}
    wb.add_sheet_data(data2)
    print wb.getData(u'新闻')
